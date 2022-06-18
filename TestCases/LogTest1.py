import time
import random
import string
import pytest
from selenium.webdriver.common.by import By
from selenium import webdriver
import openpyxl


class TestFB:
    baseURL = "https://www.facebook.com/"
    path = "C:/Users/PR PC/PycharmProjects/Facebook_Login/data.xlsx"
    wb = openpyxl.load_workbook(path)
    sht = wb['Sheet1']
    r = 2

    def test_login(self):
        self.driver = webdriver.Chrome()
        self.driver.get(self.baseURL)

        for i in range(4):
            def random_string(size=8, chars=string.ascii_lowercase + string.digits):
                return ''.join(random.choice(chars) for x in range(size))

            val = random_string(8)
            print(val)
            self.sht.cell(self.r, 1).value = val
            self.sht.cell(self.r, 2).value = val
            self.r = self.r + 1
            self.wb.save(self.path)
            self.row = self.sht.max_row
        for j in range(2, self.row + 1):
            user = self.sht.cell(j, 1).value
            pasd = self.sht.cell(j, 2).value
            self.driver.find_element(By.ID, "email").send_keys(user)
            self.driver.find_element(By.ID, "pass").send_keys(pasd)
            self.driver.find_element(By.NAME, "login").click()
            time.sleep(3)
        self.driver.quit()


if __name__ == "__main__":
    pytest.main()
